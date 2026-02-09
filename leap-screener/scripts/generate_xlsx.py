#!/usr/bin/env python3
"""
LEAP Screener XLSX 报告生成器

读取 screener 的 markdown 报告，生成格式化 xlsx。
也可独立使用，直接从 screener 输出生成。

用法：
  python generate_xlsx.py <input_md> <output_xlsx>
  python generate_xlsx.py reports/leap-fullscan-2026-02-08.md reports/LEAP_Screener_2026-02-08.xlsx
"""

import re, sys, os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ===== 样式定义 =====
BLUE = "1A73E8"; WHITE = "FFFFFF"; DARK = "202124"; GRAY = "5F6368"
LIGHT_GRAY = "F1F3F4"; GREEN = "137333"; RED = "C5221F"; ORANGE = "E37400"
YELLOW_BG = "FFF3CD"

hfont = Font(name="Arial", bold=True, color=WHITE, size=11)
hfill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
dfont = Font(name="Arial", color=DARK, size=10)
bfont = Font(name="Arial", bold=True, color=DARK, size=11)
flagfill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
altfill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
border = Border(
    left=Side("thin", color="DADCE0"), right=Side("thin", color="DADCE0"),
    top=Side("thin", color="DADCE0"), bottom=Side("thin", color="DADCE0"))
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont; cell.fill = hfill; cell.alignment = center; cell.border = border
    ws.row_dimensions[row].height = 28


def style_row(ws, row, ncols, alt=False, flag=False):
    fill = flagfill if flag else (altfill if alt else None)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = dfont; cell.border = border; cell.alignment = center
        if fill:
            cell.fill = fill
    ws.cell(row=row, column=2).font = bfont
    ws.cell(row=row, column=2).alignment = center


def parse_report(md_path):
    """从 screener markdown 报告中提取数据"""
    with open(md_path, "r") as f:
        text = f.read()

    stocks = []
    flagged = []

    # 分割 clean / flagged 区域
    clean_section = ""
    flag_section = ""

    if "无异常标记" in text:
        parts = text.split("⚠️")
        clean_section = parts[0] if len(parts) > 0 else text
        flag_section = parts[1] if len(parts) > 1 else ""
    else:
        clean_section = text

    def extract_stocks(section):
        results = []
        blocks = re.split(r'###\s*#?\d+', section)
        for block in blocks[1:]:  # skip first empty
            s = {}
            # ticker and score
            m = re.search(r'(\w+)\s*[|｜]\s*综合评分[：:]\s*\*?\*?(\d+\.?\d*)', block)
            if not m:
                m = re.search(r'(\w+)\s*—?\s*(\d+\.?\d*)\s*分', block)
            if m:
                s['ticker'] = m.group(1)
                s['score'] = float(m.group(2))
            else:
                continue

            # price
            m = re.search(r'(?:当前)?价格[：:]\s*\$?([\d,]+\.?\d*)', block)
            s['price'] = float(m.group(1).replace(',', '')) if m else 0

            # market cap
            m = re.search(r'市值\s*\$?([\d,.]+)B', block)
            s['mcap'] = m.group(1) if m else ""

            # PEG
            m = re.search(r'PEG\s*([\d.]+)', block)
            s['peg'] = float(m.group(1)) if m else 0

            # revenue growth
            m = re.search(r'收入YoY\s*([\d.]+)%', block)
            s['rev'] = m.group(1) + "%" if m else ""

            # EPS growth
            m = re.search(r'EPS增速[^F]*?([\d.-]+)%', block)
            s['eps'] = m.group(1) + "%" if m else ""

            # Forward EPS
            m = re.search(r'(?:Forward|Fwd)\s*(?:EPS)?(?:增速)?\s*([\d.]+)%', block)
            s['fwd_eps'] = m.group(1) + "%" if m else ""

            # RSI
            m = re.search(r'RSI\(?14\)?\s*([\d.]+)', block)
            s['rsi'] = float(m.group(1)) if m else 0

            # drawdown
            m = re.search(r'距52周高\s*(-?[\d.]+)%', block)
            s['dd52'] = m.group(1) + "%" if m else ""

            # upside
            m = re.search(r'Upside\s*([\d.]+)%', block)
            s['upside'] = m.group(1) + "%" if m else ""

            # LEAP
            m = re.search(r'(?:到期日|LEAP)[：:]*\s*\*?\*?([\d/.-]+)\*?\*?.*?Strike\s*\*?\*?\$?([\d,.]+)', block)
            if m:
                s['leap'] = f"${m.group(2)}C {m.group(1)}"
            else:
                m = re.search(r'LEAP[：:]*\s*\$?([\d,.]+C\s*[\d/]+)', block)
                s['leap'] = f"${m.group(1)}" if m else ""

            # IV Rank
            m = re.search(r'IV Rank[^0-9]*([\d.]+)', block)
            s['iv_rank'] = m.group(1) if m else ""

            # spread
            m = re.search(r'点差\s*([\d.]+)%', block)
            s['spread'] = m.group(1) + "%" if m else ""

            # required move
            m = re.search(r'翻倍所?需涨[幅]?\s*([\d.]+)%', block)
            s['move2x'] = m.group(1) + "%" if m else ""

            # flags
            m = re.search(r'⚠️\s*(?:异常标记|异常)[：:]\s*(.+?)(?:\n|$)', block)
            s['flag'] = m.group(1).strip() if m else ""

            results.append(s)
        return results

    stocks = extract_stocks(clean_section)
    flagged = extract_stocks(flag_section)

    # 提取元数据
    meta = {}
    m = re.search(r'生成时间[：:]\s*(.+?)$', text, re.M)
    meta['time'] = m.group(1).strip() if m else datetime.now().strftime('%Y-%m-%d %H:%M')
    m = re.search(r'扫描股票数[：:]\s*(\d+)', text)
    meta['total'] = m.group(1) if m else "517"
    m = re.search(r'全部通过.*?(\d+)', text)
    meta['passed'] = m.group(1) if m else str(len(stocks) + len(flagged))

    return stocks, flagged, meta


def build_xlsx(stocks, flagged, meta, output_path):
    wb = Workbook()

    # ===== Sheet 1: Top 15 =====
    ws = wb.active
    ws.title = "Top 15"
    ws.sheet_properties.tabColor = BLUE

    headers = ["#", "Ticker", "评分", "价格", "市值", "PEG", "收入YoY", "EPS增速",
               "Fwd EPS增速", "RSI", "距52周高", "Upside", "LEAP合约", "IV Rank", "点差", "翻倍需涨"]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, s in enumerate(stocks[:15]):
        r = 2 + i
        vals = [i + 1, s['ticker'], s['score'], s['price'], s.get('mcap', ''),
                s['peg'], s.get('rev', ''), s.get('eps', ''), s.get('fwd_eps', ''),
                s['rsi'], s.get('dd52', ''), s.get('upside', ''), s.get('leap', ''),
                s.get('iv_rank', ''), s.get('spread', ''), s.get('move2x', '')]
        for c, val in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), alt=(i % 2 == 1))

    widths = [4, 8, 7, 11, 9, 6, 9, 9, 10, 7, 9, 8, 14, 8, 7, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "C2"

    # Color scales
    n = min(len(stocks), 15)
    if n > 1:
        for col, reverse in [("C", False), ("J", True), ("P", True), ("F", True)]:
            rng = f"{col}2:{col}{1 + n}"
            if reverse:
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="min", start_color="D4EDDA",
                    mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                    end_type="max", end_color="F8D7DA"))
            else:
                ws.conditional_formatting.add(rng, ColorScaleRule(
                    start_type="min", start_color="F8D7DA",
                    mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                    end_type="max", end_color="D4EDDA"))

    # ===== Sheet 2: Flagged =====
    ws2 = wb.create_sheet("⚠️ 有Flag")
    ws2.sheet_properties.tabColor = ORANGE

    fheaders = ["#", "Ticker", "评分", "价格", "PEG", "RSI", "翻倍需涨", "LEAP合约", "⚠️ 异常原因"]
    for c, h in enumerate(fheaders, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(fheaders))

    for i, s in enumerate(flagged):
        r = 2 + i
        vals = [i + 1, s['ticker'], s['score'], s['price'], s['peg'],
                s['rsi'], s.get('move2x', ''), s.get('leap', ''), s.get('flag', '')]
        for c, val in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=val)
        style_row(ws2, r, len(fheaders), flag=True)
        ws2.cell(row=r, column=9).alignment = left_align

    fwidths = [4, 8, 7, 11, 6, 7, 10, 14, 40]
    for c, w in enumerate(fwidths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "C2"

    # ===== Sheet 3: 观察 =====
    ws3 = wb.create_sheet("💡 观察")
    ws3.sheet_properties.tabColor = GREEN
    observations = [
        ["📐 指标说明", ""],
        ["PEG", "市盈率/增长率比，<1 低估"],
        ["RSI", "相对强弱，<30 超卖"],
        ["IV Rank", "隐含波动率历史位置（0=最低 100=最高）"],
        ["翻倍需涨", "股价需涨多少让 LEAP Call 翻倍"],
        ["", ""],
        ["🔍 筛选条件", ""],
        ["Layer 1", "收入>10% · EPS>15% · PEG<2 · 市值>$5B · Upside>15%"],
        ["Layer 2", "RSI<60 · 距52周高>-35%"],
        ["Layer 3", "LEAP>9月 · ATM点差 · IV Rank · 翻倍涨幅"],
        ["", ""],
        [f"扫描 {meta.get('total', '517')} 只", f"通过 {meta.get('passed', '?')} 只"],
        ["生成时间", meta.get('time', '')],
    ]
    for i, (a, b) in enumerate(observations):
        r = 1 + i
        ws3.cell(row=r, column=1, value=a).font = bfont
        ws3.cell(row=r, column=2, value=b).font = dfont
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 55

    wb.save(output_path)
    print(f"OK: {output_path}")


def main():
    if len(sys.argv) < 3:
        print(f"用法: {sys.argv[0]} <input.md> <output.xlsx>")
        sys.exit(1)

    stocks, flagged, meta = parse_report(sys.argv[1])
    print(f"解析完成: {len(stocks)} clean + {len(flagged)} flagged")
    build_xlsx(stocks, flagged, meta, sys.argv[2])


if __name__ == "__main__":
    main()
